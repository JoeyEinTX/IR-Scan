import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Read the floor report CSV
with open('Equipment_Report_By_Floor.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    data = list(reader)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Equipment by Floor"

# Define styles
title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')

floor_header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
floor_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

column_header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
column_header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')

summary_font = Font(name='Calibri', size=10, bold=True, color='000000')
summary_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

grand_total_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
grand_total_fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')

normal_font = Font(name='Calibri', size=10)
alternate_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Write data with formatting
row_num = 1
in_data_section = False
data_row_counter = 0

for csv_row in data:
    if not csv_row or all(cell == '' for cell in csv_row):
        row_num += 1
        continue
    
    # Check row type and apply formatting
    row_text = ' '.join(csv_row).strip()
    
    # Title rows
    if 'HOSPITAL ELECTRICAL EQUIPMENT' in row_text or 'Organized by Floor' in row_text:
        merged_cell = f'A{row_num}:J{row_num}'
        ws.merge_cells(merged_cell)
        cell = ws[f'A{row_num}']
        cell.value = csv_row[0]
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = center_alignment
        ws.row_dimensions[row_num].height = 25
    
    # Floor headers
    elif row_text.startswith('FLOOR:'):
        merged_cell = f'A{row_num}:J{row_num}'
        ws.merge_cells(merged_cell)
        cell = ws[f'A{row_num}']
        cell.value = row_text
        cell.font = floor_header_font
        cell.fill = floor_header_fill
        cell.alignment = center_alignment
        ws.row_dimensions[row_num].height = 22
        in_data_section = True
        data_row_counter = 0
    
    # Column headers
    elif 'Room Number' in row_text and 'Equipment Type' in row_text:
        for col_num, value in enumerate(csv_row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = column_header_font
            cell.fill = column_header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        ws.row_dimensions[row_num].height = 20
    
    # Summary sections
    elif 'FLOOR SUMMARY' in row_text or 'FLOOR TOTAL' in row_text:
        for col_num, value in enumerate(csv_row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = summary_font
            cell.fill = summary_fill
            cell.alignment = left_alignment if col_num <= 2 else center_alignment
            cell.border = thin_border
        in_data_section = False
    
    # Grand total section
    elif 'GRAND TOTAL' in row_text or '======' in row_text:
        if '======' not in row_text:
            merged_cell = f'A{row_num}:J{row_num}'
            ws.merge_cells(merged_cell)
            cell = ws[f'A{row_num}']
            cell.value = row_text
            cell.font = grand_total_font
            cell.fill = grand_total_fill
            cell.alignment = center_alignment
            ws.row_dimensions[row_num].height = 22
    
    # Equipment Type breakdown in grand total
    elif any(equip_type in row_text for equip_type in ['Automatic Transfer', 'Branch Panel', 'Disconnect', 'ECB', 'Feeder', 'LIM Panel', 'Main Breaker', 'Transformer']):
        for col_num, value in enumerate(csv_row, 1):
            if value:
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.font = normal_font
                cell.fill = summary_fill
                cell.alignment = left_alignment if col_num == 1 else center_alignment
                cell.border = thin_border
    
    # Regular data rows
    elif in_data_section and csv_row[0] and csv_row[0] not in ['', 'FLOOR SUMMARY']:
        # Alternate row colors
        fill = alternate_fill if data_row_counter % 2 == 1 else PatternFill()
        
        for col_num, value in enumerate(csv_row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = normal_font
            cell.fill = fill
            cell.alignment = left_alignment if col_num <= 3 else center_alignment
            cell.border = thin_border
        
        data_row_counter += 1
    
    row_num += 1

# Auto-adjust column widths
column_widths = {
    'A': 25,  # Room Number
    'B': 20,  # Equipment Type
    'C': 35,  # Equipment Name/Label
    'D': 15,  # Amperage/kVA
    'E': 18,  # Voltage
    'F': 12,  # Phase/Wire
    'G': 20,  # Fed From
    'H': 15,  # Mounting
    'I': 15,  # System Type
    'J': 40   # Notes
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Freeze top rows
ws.freeze_panes = 'A4'

# Save the workbook
wb.save('Equipment_Report_By_Floor.xlsx')
print("Created: Equipment_Report_By_Floor.xlsx")
print("\nFormatting applied:")
print("  - Color-coded headers and sections")
print("  - Alternating row colors for readability")
print("  - Auto-sized columns")
print("  - Professional fonts and styling")
print("  - Borders and alignment")
print("\nReady to open in Excel!")

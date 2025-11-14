import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Read the consolidated master file to get panel information
panels_data = []
with open('Consolidated_Hospital_Equipment_Master.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        if row['Equipment Type'] in ['Branch Panel', 'Feeder Panel']:
            panels_data.append(row)

# Sort by Floor and Room
floor_order = {
    'TEP Plant': -1,
    'Ground': 0,
    '2nd Floor': 2,
    '3rd Floor': 3,
    '4th Floor': 4,
    '5th Floor': 5,
    '6th Floor': 6,
    '7th Floor': 7,
    '8th Floor': 8,
    '9th Floor': 9
}

panels_data.sort(key=lambda x: (floor_order.get(x['Floor'], 999), x['Room Number']))

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Panel Inspection"

# Define styles
header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Add headers matching the required format
headers = ['PANEL NUMBER', 'LOCATION', 'LOCKED', 'CLEAN', 'DAY', 'TECH', 'NOTES']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

ws.row_dimensions[1].height = 25

# Add data rows
for row_num, panel in enumerate(panels_data, 2):
    # Column A: PANEL NUMBER (Panel Name/Label)
    cell = ws.cell(row=row_num, column=1)
    cell.value = panel['Equipment Name/Label']
    cell.border = border
    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Column B: LOCATION (Floor - Room Number)
    cell = ws.cell(row=row_num, column=2)
    cell.value = f"{panel['Floor']} - {panel['Room Number']}"
    cell.border = border
    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Columns C-G: Empty for inspection data (LOCKED, CLEAN, DAY, TECH, NOTES)
    for col_num in range(3, 8):
        cell = ws.cell(row=row_num, column=col_num)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Alternate row colors for readability
    if row_num % 2 == 0:
        for col_num in range(1, 8):
            ws.cell(row=row_num, column=col_num).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

# Adjust column widths
column_widths = {
    'A': 20,  # PANEL NUMBER
    'B': 30,  # LOCATION
    'C': 12,  # LOCKED
    'D': 12,  # CLEAN
    'E': 12,  # DAY
    'F': 12,  # TECH
    'G': 30   # NOTES
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Freeze the header row
ws.freeze_panes = 'A2'

# Save workbook
wb.save('Panel_Inspection_Report.xlsx')
print("✓ Panel_Inspection_Report.xlsx created successfully")
print(f"  - Total panels: {len(panels_data)}")
print("  - Columns: PANEL NUMBER | LOCATION | LOCKED | CLEAN | DAY | TECH | NOTES")
